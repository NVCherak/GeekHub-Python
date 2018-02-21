import requests
from bs4 import BeautifulSoup as BS
import os
import argparse
import re
import json
import csv
from openpyxl import Workbook


cmdParser = argparse.ArgumentParser(description='Select parser settings')
# adding up arguments for command prompt
cmdParser.add_argument('-g', default='quotes', choices=['quotes', 'authors', 'tags', '.'])
cmdParser.add_argument('-f', default='txt', choices=['txt', 'csv', 'json', 'xls', '.'])
cmdParser.add_argument('-a', metavar='N', type=int, nargs='+')

print('In progress...')

folder = './results/'
directory = os.path.dirname(folder)
if not os.path.exists(directory):
    # checking if exist the folder and create if it's not exist
    makedirs = os.makedirs(directory)

url = 'http://quotes.toscrape.com'  # url address of site
try:
    # try to connect with it
    page = requests.get(url)
except Exception as e:
    print('Error!' + str(e))
soup = BS(page.content, 'html.parser')  # parsing the page


def get_all_quotes(parser, quotes):
    # function getting all record of quotes from the site
    for soupQuote in parser.find_all('div', 'quote'):
        quote = {'text': None, 'author': None, 'tags': []}

        quote['text'] = soupQuote.find('span', 'text').text[1:-2]
        quote['author'] = soupQuote.find('small', 'author').text
        for tag in soupQuote.find_all('a', 'tag'):
            quote['tags'].append(tag.text)

        quotes.append(quote)

    if parser.find('li', 'next'):
        next_url = parser.find('li', 'next').find('a').attrs['href']
        try:
            next_page = requests.get(url + next_url)
        except Exception as e:
            print('Error!' + str(e))
        new_soup = BS(next_page.content, 'html.parser')
        get_all_quotes(new_soup, quotes)

    return quotes


def get_all_authors(parser, authors):
    # function getting all record of authors from the site
    for soupQuote in parser.find_all('div', 'quote'):
        author = {'author-title': None, 'url': None, 'born_date': None, 'about': None}

        url_to_author = soupQuote.find('a').attrs['href']
        new_url = url + url_to_author

        if not [item for item in authors if item['url'] == new_url]:
            try:
                author_page = requests.get(new_url)
            except Exception as e:
                print('Error!' + str(e))

            author['url'] = new_url
            author_soup = BS(author_page.content, 'html.parser')
            author['author-title'] = author_soup.find('h3', 'author-title').text
            author['born_date'] = author_soup.find('span', 'author-born-date').text + ' ' + \
                                  author_soup.find('span', 'author-born-location').text
            author['about'] = re.sub(r'(^\n( )*)|(( )*\n( )*$)|\n', '',
                                     author_soup.find('div', 'author-description').text)
            authors.append(author)


    if parser.find('li', 'next'):
        next_url = parser.find('li', 'next').find('a').attrs['href']
        try:
            next_page = requests.get(url + next_url)
        except Exception as e:
            print('Error!' + str(e))
        new_soup = BS(next_page.content, 'html.parser')
        get_all_authors(new_soup, authors)

    return authors


def get_authors_by_id(parser, authors, listId, innerId=0):
    # function getting record of authors by id from the site
    for soupQuote in parser.find_all('div', 'quote'):
        author = {'author-title': None, 'url': None, 'born_date': None, 'about': None}

        url_to_author = soupQuote.find('a').attrs['href']
        new_url = url + url_to_author

        if not [item for item in authors if item['url'] == new_url]:
            innerId += 1
            if innerId in listId:
                try:
                    author_page = requests.get(new_url)
                except Exception as e:
                    print('Error!' + str(e))
                author['url'] = new_url
                author_soup = BS(author_page.content, 'html.parser')
                author['author-title'] = author_soup.find('h3', 'author-title').text
                author['born_date'] = author_soup.find('span', 'author-born-date').text + ' ' + \
                                      author_soup.find('span', 'author-born-location').text
                author['about'] = re.sub(r'(^\n( )*)|(( )*\n( )*$)|\n', '',
                                         author_soup.find('div', 'author-description').text)
                authors.append(author)

                listId.remove(innerId)
                if not listId:
                    return authors

    if parser.find('li', 'next'):
        next_url = parser.find('li', 'next').find('a').attrs['href']
        try:
            next_page = requests.get(url + next_url)
        except Exception as e:
            print('Error!' + str(e))
        new_soup = BS(next_page.content, 'html.parser')
        get_authors_by_id(new_soup, authors, listId, innerId)

    return authors

def get_all_tags(parser, tags):
    # function getting all record of tags from site
    for soupTag in parser.find_all('a', 'tag'):
        tag = {'tag-title': None, 'url': None, 'authors': [], 'authors_url': []}

        if not [item for item in tags if item['tag-title'] == soupTag.text]:
            tag['tag-title'] = soupTag.text
            tag['url'] = url + soupTag.attrs['href']

            try:
                tag_page = requests.get(tag['url'])
            except Exception as e:
                print('Error!' + str(e))
            soup_tag_page = BS(tag_page.content, 'html.parser')

            for soup_tag_quote in soup_tag_page.find_all('div', 'quote'):
                if not [item for item in tag['authors'] if item == soup_tag_quote.find('small', 'author').text]:
                    tag['authors'].append(soup_tag_quote.find('small', 'author').text)
                    tag['authors_url'].append(url + soup_tag_quote.find('a').attrs['href'])

            while soup_tag_page.find('li', 'next'):
                next_url = soup_tag_page.find('li', 'next').find('a').attrs['href']
                try:
                    next_page = requests.get(url + next_url)
                except Exception as e:
                    print('Error!' + str(e))
                soup_tag_page = BS(next_page.content, 'html.parser')
                for soup_tag_quote in soup_tag_page.find_all('div', 'quote'):
                    if not [item for item in tag['authors'] if item == soup_tag_quote.find('small', 'author').text]:
                        tag['authors'].append(soup_tag_quote.find('small', 'author').text)
                        tag['authors_url'].append(url + soup_tag_quote.find('a').attrs['href'])

            tags.append(tag)

    if parser.find('li', 'next'):
        next_url = parser.find('li', 'next').find('a').attrs['href']
        try:
            next_page = requests.get(url + next_url)
        except Exception as e:
            print('Error!' + str(e))
        new_soup = BS(next_page.content, 'html.parser')
        get_all_tags(new_soup, tags)

    return tags


def write_quotes_to_file(extension=cmdParser.parse_args().f):
    # function writing quotes records to file (with the option of format select)
    parsing_list = get_all_quotes(soup, [])

    def write_quotes_to_txt():
        with open(folder + 'quotes.txt', 'w', encoding='utf-8') as txtfile:
            count = 0
            for quote in parsing_list:
                count += 1
                record = str(count) + '\ttext: ' + quote['text'] + '\n\tauthor: ' + quote['author'] + '\n\ttags: '
                for tag in quote['tags']:
                    record += tag + ', '
                record = re.sub(r', $', '\n', record)
                txtfile.write(record)

    def write_quotes_to_csv():
        with open(folder + 'quotes.csv', 'w', encoding='utf-8') as csvfile:
            writer = csv.DictWriter(csvfile, fieldnames=parsing_list[0].keys())
            writer.writeheader()
            for DictRecord in parsing_list:
                writer.writerow(DictRecord)

    def write_quotes_to_json():
        with open(folder + 'quotes.json', 'w', encoding='utf-8') as jsonfile:
            for record in parsing_list:
                json.dump(record, jsonfile, sort_keys=True, indent=4)

    def write_quotes_to_xls():
        wb = Workbook()
        ws = wb.active
        for i, record in zip(range(len(parsing_list)), parsing_list):
            ws['A' + str(i + 1)] = record['text']
            ws['B' + str(i + 1)] = record['author']
            ws['C' + str(i + 1)] = re.sub(r"[\[\]']", '', str(record['tags']))
        wb.save(folder + 'quotes.xls')

    if extension == 'txt':
        write_quotes_to_txt()
    elif extension == 'csv':
        write_quotes_to_csv()
    elif extension == 'json':
        write_quotes_to_json()
    elif extension == 'xls':
        write_quotes_to_xls()
    elif extension == '.':
        write_quotes_to_txt()
        write_quotes_to_csv()
        write_quotes_to_json()
        write_quotes_to_xls()


def write_authors_to_file(parsing_list=get_all_authors(soup, []), extension=cmdParser.parse_args().f):
    # function writing authors records to file (with the option of format selection)
    def write_authors_to_txt():
        with open(folder + 'authors.txt', 'w', encoding='utf-8') as txtfile:
            count = 0
            for author in parsing_list:
                count += 1
                record = str(count) + '\turl: ' + author['url'] + '\n\tauthor: ' + author['author-title'] + \
                         '\n\tborn_date: ' + author['born_date'] + '\n\tabout: ' + author['about'] + '\n'
                txtfile.write(record)

    def write_authors_to_csv():
        with open(folder + 'authors.csv', 'w', encoding='utf-8') as csvfile:
            writer = csv.DictWriter(csvfile, fieldnames=parsing_list[0].keys())
            writer.writeheader()
            for DictRecord in parsing_list:
                writer.writerow(DictRecord)

    def write_authors_to_json():
        with open(folder + 'authors.json', 'w', encoding='utf-8') as jsonfile:
            for record in parsing_list:
                json.dump(record, jsonfile, sort_keys=True, indent=4)

    def write_authors_to_xls():
        wb = Workbook()
        ws = wb.active
        for i, record in zip(range(len(parsing_list)), parsing_list):
            ws['B' + str(i + 1)] = record['author-title']
            ws['A' + str(i + 1)] = record['url']
            ws['C' + str(i + 1)] = record['born_date']
            ws['D' + str(i + 1)] = record['about']
        wb.save(folder + 'authors.xls')

    if extension == 'txt':
        write_authors_to_txt()
    elif extension == 'csv':
        write_authors_to_csv()
    elif extension == 'json':
        write_authors_to_json()
    elif extension == 'xls':
        write_authors_to_xls()
    elif extension == '.':
        write_authors_to_txt()
        write_authors_to_csv()
        write_authors_to_json()
        write_authors_to_xls()


def write_authors_by_id_to_file(listId=cmdParser.parse_args().a):
    # function writing authors records by id to file (with the option of format select)
    parsing_list = get_authors_by_id(soup, [], listId)
    write_authors_to_file(parsing_list)

def write_tags_to_file(extension=cmdParser.parse_args().f):
    # function writing tags records to file (with the option of format select)
    parsing_list = get_all_tags(soup, [])

    def write_tags_to_txt():
        with open(folder + 'tags.txt', 'w', encoding='utf-8') as txtfile:
            count = 0
            for tag in parsing_list:
                count += 1
                record = str(count) + '\ttag title: ' + tag['tag-title'] + '\n\turl: ' + tag['url'] + '\n\tauthors: '
                for author in tag['authors']:
                    record += author + ', '
                record = re.sub(r', $', '\n', record)
                record += '\tauthors_url: '
                for author_url in tag['authors_url']:
                    record += author_url + ', '
                record = re.sub(r', $', '\n', record)

                txtfile.write(record)

    def write_tags_to_csv():
        with open(folder + 'tags.csv', 'w', encoding='utf-8') as csvfile:
            writer = csv.DictWriter(csvfile, fieldnames=parsing_list[0].keys())
            writer.writeheader()
            for DictRecord in parsing_list:
                writer.writerow(DictRecord)

    def write_tags_to_json():
        with open(folder + 'tags.json', 'w', encoding='utf-8') as jsonfile:
            for record in parsing_list:
                json.dump(record, jsonfile, sort_keys=True, indent=4)

    def write_tags_to_xls():
        wb = Workbook()
        ws = wb.active
        for i, record in zip(range(len(parsing_list)), parsing_list):
            ws['A' + str(i + 1)] = record['url']
            ws['B' + str(i + 1)] = record['tag-title']
            ws['C' + str(i + 1)] = re.sub(r"[\[\]']", '', str(record['authors']))
            ws['D' + str(i + 1)] = re.sub(r"[\[\]']", '', str(record['authors_url']))
        wb.save(folder + 'tags.xls')

    if extension == 'txt':
        write_tags_to_txt()
    elif extension == 'csv':
        write_tags_to_csv()
    elif extension == 'json':
        write_tags_to_json()
    elif extension == 'xls':
        write_tags_to_xls()
    elif extension == '.':
        write_tags_to_txt()
        write_tags_to_csv()
        write_tags_to_json()
        write_tags_to_xls()


if cmdParser.parse_args().a:
    # check comand prompt arguments
    write_authors_by_id_to_file()
else:
    if cmdParser.parse_args().g == 'quotes':
        write_quotes_to_file()
    elif cmdParser.parse_args().g == 'authors':
        write_authors_to_file()
    elif cmdParser.parse_args().g == 'tags':
        write_tags_to_file()
    elif cmdParser.parse_args().g == '.':
        write_quotes_to_file()
        write_authors_to_file()
        write_tags_to_file()

print('Complete!')
